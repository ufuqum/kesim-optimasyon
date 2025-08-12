import json
import csv
import os
from tkinter import filedialog, messagebox
from typing import Optional, Tuple, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.units import mm
from optimization import calculate_costs


def safe_get_part_attr(part: Any, attr: str, default: Any = "") -> Any:
    """
    Parça verisinden hem NamedTuple, dict, hem de tuple/list tipleri için güvenli veri erişimi sağlar.
    """
    if hasattr(part, attr):
        return getattr(part, attr)
    if isinstance(part, dict):
        return part.get(attr, default)
    if isinstance(part, (list, tuple)):
        # Index bazında erişim örnekleri
        if attr == "name":
            return part[0] if len(part) > 0 else default
        if attr == "length":
            return part[1] if len(part) > 1 else default
        if attr == "quantity":
            return part[2] if len(part) > 2 else default
        if attr == "cut_type":
            return part[3] if len(part) > 3 else default
    return default


def save_project(parts_data: List[Any], stock_length: int, kerf: int) -> None:
    file_path = filedialog.asksaveasfilename(
        defaultextension=".json",
        filetypes=[("JSON Dosyaları", "*.json"), ("Tüm Dosyalar", "*.*")]
    )
    if not file_path:
        return
    data = {
        "parts": parts_data,
        "stock_length": stock_length,
        "kerf": kerf
    }
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        messagebox.showinfo("Başarılı", "Proje başarıyla kaydedildi.")
    except Exception as e:
        messagebox.showerror("Hata", f"Dosya kaydedilirken hata oluştu:\n{e}")


def load_project() -> Optional[Tuple[List[Any], int, int]]:
    file_path = filedialog.askopenfilename(
        filetypes=[("JSON Dosyaları", "*.json"), ("Tüm Dosyalar", "*.*")]
    )
    if not file_path or not os.path.exists(file_path):
        return None
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        parts = data.get("parts", [])
        stock_length = data.get("stock_length", 6000)
        kerf = data.get("kerf", 3)
        return parts, stock_length, kerf
    except Exception as e:
        messagebox.showerror("Hata", f"Dosya yüklenirken hata oluştu:\n{e}")
        return None


def export_to_csv(parts_data: List[Any]) -> None:
    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV Dosyaları", "*.csv")]
    )
    if not file_path:
        return
    try:
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Parça Adı", "Uzunluk (mm)", "Adet"])
            for part in parts_data:
                writer.writerow([
                    safe_get_part_attr(part, "name", ""),
                    safe_get_part_attr(part, "length", ""),
                    safe_get_part_attr(part, "quantity", "")
                ])
        messagebox.showinfo("Başarılı", "CSV dosyası başarıyla oluşturuldu.")
    except Exception as e:
        messagebox.showerror("Hata", f"CSV dışa aktarılırken hata oluştu:\n{e}")


def import_from_csv() -> Optional[List[Any]]:
    file_path = filedialog.askopenfilename(filetypes=[("CSV Dosyaları", "*.csv")])
    if not file_path:
        return None
    imported = []
    try:
        with open(file_path, newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                try:
                    length = float(row.get("Uzunluk (mm)", 0))
                    quantity = int(row.get("Adet", 0))
                    name = row.get("Parça Adı", "")
                    if length > 0 and quantity > 0:
                        imported.append({"name": name, "length": length, "quantity": quantity})
                except ValueError:
                    continue
        return imported
    except Exception as e:
        messagebox.showerror("Hata", f"CSV içe aktarılırken hata oluştu:\n{e}")
        return None


def export_to_excel(optimization_result: dict, file_path: str, stock_unit_price: float) -> None:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Kesim Planı"
        bold_font = Font(bold=True)

        ws.append(["Kesim Planı Raporu"])
        ws.append([])

        plan = optimization_result.get("plan", [])
        fire_eff = optimization_result.get("fire_efficiency", {})
        costs = calculate_costs(fire_eff, stock_unit_price)

        # Stok bazlı plan
        for stok_no, stock_parts in enumerate(plan, start=1):
            ws.append([f"Stok {stok_no}"])

            # Parçaları grupla ve adetleri topla
            part_counts = {}
            for part in stock_parts:
                p_name = safe_get_part_attr(part, "name", "")
                p_length_raw = safe_get_part_attr(part, "length", 0)
                p_cut_type = safe_get_part_attr(part, "cut_type", "")
                try:
                    p_length = float(p_length_raw)
                except (TypeError, ValueError):
                    p_length = 0.0
                key = (p_name, p_length, p_cut_type)
                part_counts[key] = part_counts.get(key, 0) + 1

            # Grupları "Adet x Uzunluk" formatında metne dönüştür
            stok_desc_list = []
            for (p_name, p_length, p_cut_type), count in part_counts.items():
                if p_cut_type and p_cut_type.lower() != "düz kesim":
                    stok_desc_list.append(f"{p_name} ({p_length:.1f} mm, {p_cut_type}) x {count}")
                else:
                    stok_desc_list.append(f"{p_name} ({p_length:.1f} mm) x {count}")

            stok_desc = " - ".join(stok_desc_list)
            ws.append([stok_desc])
            ws.append([])  # stoklar arası boşluk

        # Genel istatistikler
        ws.append([])
        ws.append(["Toplam Fire (mm):", fire_eff.get("total_fire", 0)])
        ws.append(["Toplam Verimlilik (%):", fire_eff.get("total_efficiency", 0)])
        ws.append(["Toplam Maliyet:", costs.get("total_cost", 0)])
        ws.append(["Fire Maliyeti:", costs.get("fire_cost", 0)])

        # Parça listesi başlığı
        ws.append([])
        ws.append(["Parça Listesi:"])
        ws.append(["Parça Adı", "Uzunluk (mm)", "Adet", "Kesim Tipi"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        parts_list = optimization_result.get("parts_list", [])
        if not parts_list:
            part_aggregate = {}
            for stk in plan:
                for part in stk:
                    p_name = safe_get_part_attr(part, "name", "")
                    p_length_raw = safe_get_part_attr(part, "length", 0)
                    try:
                        p_length = float(p_length_raw)
                    except (TypeError, ValueError):
                        p_length = 0.0
                    key = (p_name, p_length)
                    part_aggregate[key] = part_aggregate.get(key, 0) + 1
            parts_list = [(k[0], k[1], v) for k, v in part_aggregate.items()]

        for part in parts_list:
            p_name = safe_get_part_attr(part, "name", "")
            p_len_raw = safe_get_part_attr(part, "length", 0)
            p_qty = safe_get_part_attr(part, "quantity", 0)
            p_cut_type = safe_get_part_attr(part, "cut_type", "")

            try:
                p_len = float(p_len_raw)
            except (TypeError, ValueError):
                p_len = 0.0

            if p_cut_type and p_cut_type.lower() != "düz kesim":
                ws.append([p_name, p_len, p_qty, p_cut_type])
            else:
                ws.append([p_name, p_len, p_qty])

        wb.save(file_path)
        messagebox.showinfo("Başarılı", f"Excel dosyası başarıyla oluşturuldu:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Hata", f"Excel dışa aktarılırken hata oluştu:\n{e}")


def export_to_pdf(optimization_result: dict, file_path: str, stock_unit_price: float) -> None:
    try:
        c = pdfcanvas.Canvas(file_path, pagesize=letter)
        width, height = letter
        margin = 20 * mm
        y = height - margin

        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin, y, "Kesim Planİ Raporu")
        y -= 18
        c.setFont("Helvetica", 10)

        plan = optimization_result.get("plan", [])
        fire_eff = optimization_result.get("fire_efficiency", {})
        costs = calculate_costs(fire_eff, stock_unit_price)

        # Stok bazlı plan ve parça listesi
        for stok_no, stock_parts in enumerate(plan, start=1):
            c.drawString(margin, y, f"Stok {stok_no}:")
            y -= 14

            part_counts = {}
            for part in stock_parts:
                p_name = safe_get_part_attr(part, "name", "")
                p_length_raw = safe_get_part_attr(part, "length", 0)
                p_cut_type = safe_get_part_attr(part, "cut_type", "")
                try:
                    p_length = float(p_length_raw)
                except (TypeError, ValueError):
                    p_length = 0.0
                key = (p_name, p_length, p_cut_type)
                part_counts[key] = part_counts.get(key, 0) + 1

            for (p_name, p_length, p_cut_type), count in part_counts.items():
                line = f"- {p_name} ({p_length:.1f} mm) x {count}"
                if p_cut_type and p_cut_type.lower() != "düz kesim":
                    line = f"- {p_name} ({p_length:.1f} mm, {p_cut_type}) x {count}"

                c.drawString(margin + 10, y, line)
                y -= 14
                if y < margin:
                    c.showPage()
                    y = height - margin
                    c.setFont("Helvetica", 10)

            y -= 10  # stoklar arası boşluk

        # Genel özet
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, "Toplamlar ve Maliyetler:")
        y -= 14
        c.setFont("Helvetica", 10)
        c.drawString(margin + 10, y, f"Toplam Fire (mm): {fire_eff.get('total_fire', 0):.1f}")
        y -= 14
        c.drawString(margin + 10, y, f"Toplam Verimlilik (%): {fire_eff.get('total_efficiency', 0):.1f}")
        y -= 14
        c.drawString(margin + 10, y, f"Toplam Maliyet: {costs.get('total_cost', 0):.2f} TL")
        y -= 14
        c.drawString(margin + 10, y, f"Fire Maliyeti: {costs.get('fire_cost', 0):.2f} TL")
        y -= 20

        # Parça listesi başlığı
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, "Parça Listesi:")
        y -= 15
        c.setFont("Helvetica", 10)

        parts_list = optimization_result.get("parts_list", [])
        if not parts_list:
            part_aggregate = {}
            for stk in plan:
                for part in stk:
                    p_name = safe_get_part_attr(part, "name", "")
                    p_length_raw = safe_get_part_attr(part, "length", 0)
                    try:
                        p_length = float(p_length_raw)
                    except (TypeError, ValueError):
                        p_length = 0.0
                    key = (p_name, p_length)
                    part_aggregate[key] = part_aggregate.get(key, 0) + 1
            parts_list = [(k[0], k[1], v) for k, v in part_aggregate.items()]

        for part in parts_list:
            p_name = safe_get_part_attr(part, "name", "")
            p_len_raw = safe_get_part_attr(part, "length", 0)
            p_qty = safe_get_part_attr(part, "quantity", 0)
            p_cut_type = safe_get_part_attr(part, "cut_type", "")

            try:
                p_len = float(p_len_raw)
            except (TypeError, ValueError):
                p_len = 0.0

            line = f"- {p_name}: {p_qty} adet, {p_len:.1f} mm"
            if p_cut_type:
                line += f", {p_cut_type}"

            c.drawString(margin + 10, y, line)
            y -= 14
            if y < margin:
                c.showPage()
                y = height - margin
                c.setFont("Helvetica", 10)

        c.save()
        messagebox.showinfo("Başarılı", f"PDF dosyası başarıyla oluşturuldu:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Hata", f"PDF dışa aktarılırken hata oluştu:\n{e}")
